from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import sqlite3
from datetime import datetime, date
from functools import wraps
from auth import auth
from logger import log_action
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

app = Flask(__name__)
app.secret_key = 'dalessandro2010'

app.register_blueprint(auth)


def get_db():
    conn = sqlite3.connect('demandas.db')
    conn.row_factory = sqlite3.Row
    return conn


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Faça login para continuar.', 'warning')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


@app.route('/')
@login_required
def index():
    conn = get_db()
    query = '''
        SELECT d.*, p.valor as prioridade_nome 
        FROM demandas d 
        LEFT JOIN prioridades p ON d.id_prioridade = p.id
        ORDER BY p.peso DESC, d.data_criacao ASC
        LIMIT 10
    '''
    demandas = conn.execute(query).fetchall()
    total = conn.execute('SELECT COUNT(*) FROM demandas').fetchone()[0]
    conn.close()
    return render_template('index.html', demandas=demandas, total=total, offset=10)


@app.route('/mais_demandas')
@login_required
def mais_demandas():
    offset = int(request.args.get('offset', 10))
    conn = get_db()
    query = '''
        SELECT d.*, p.valor as prioridade_nome 
        FROM demandas d 
        LEFT JOIN prioridades p ON d.id_prioridade = p.id
        ORDER BY p.peso DESC, d.data_criacao ASC
        LIMIT 10 OFFSET ?
    '''
    demandas = conn.execute(query, (offset,)).fetchall()
    conn.close()
    return render_template('linhas_demandas.html', demandas=demandas)


@app.route('/nova_demanda', methods=['GET', 'POST'])
@login_required
def nova_demanda():
    if request.method == 'POST':
        titulo = request.form['titulo']
        descricao = request.form['descricao']
        id_prio = request.form['id_prioridade']
        status = request.form.get('status', 'Aberta')
        prazo = request.form['prazo'] or None

        solicitante = session['user_nome']

        conn = get_db()
        cursor = conn.execute(
            "INSERT INTO demandas (titulo, descricao, solicitante, data_criacao, id_prioridade, status, prazo) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (titulo, descricao, solicitante, datetime.now().strftime("%d/%m/%Y %H:%M"), id_prio, status, prazo)
        )
        conn.commit()
        demanda_id = cursor.lastrowid
        conn.close()

        log_action(
            session.get('user_nome', 'sistema'),
            'CRIAR_DEMANDA',
            f'Demanda: {titulo} (id={demanda_id})',
            request.remote_addr
        )

        flash('Demanda criada!', 'success')
        return redirect('/')

    conn = get_db()
    prioridades = conn.execute('SELECT * FROM prioridades ORDER BY peso DESC').fetchall()
    conn.close()
    return render_template('nova_demanda.html', prioridades=prioridades)


@app.route('/deletar/<id>')
@login_required
def deletar(id):
    conn = get_db()
    demanda = conn.execute('SELECT titulo FROM demandas WHERE id=?', (id,)).fetchone()
    titulo = demanda['titulo'] if demanda else f'id={id}'
    conn.execute('DELETE FROM demandas WHERE id=?', (id,))
    conn.commit()
    conn.close()

    log_action(
        session.get('user_nome', 'sistema'),
        'DELETAR_DEMANDA',
        f'Demanda: {titulo} (id={id})',
        request.remote_addr
    )

    return redirect('/')


@app.route('/buscar')
@login_required
def buscar():
    termo = request.args.get('q', '').strip()
    conn = get_db()
    query = '''
        SELECT d.*, p.valor as prioridade_nome 
        FROM demandas d 
        LEFT JOIN prioridades p ON d.id_prioridade = p.id 
        WHERE d.titulo LIKE ? OR p.valor LIKE ?
        ORDER BY p.peso DESC, d.id ASC
    '''
    filtro = f'%{termo}%'
    resultados = conn.execute(query, (filtro, filtro)).fetchall()
    conn.close()
    total = len(resultados)
    return render_template('index.html', demandas=resultados, total=total, offset=total)


@app.route('/detalhes/<id>')
@login_required
def detalhes(id):
    conn = get_db()
    demanda = conn.execute('''
        SELECT d.*, p.valor as prioridade_nome 
        FROM demandas d 
        LEFT JOIN prioridades p ON d.id_prioridade = p.id 
        WHERE d.id=?''', (id,)).fetchone()
    comentarios = conn.execute('SELECT * FROM comentarios WHERE demanda_id=?', (id,)).fetchall()
    conn.close()
    return render_template('detalhes.html', demanda=demanda, comentarios=comentarios)


@app.route('/adicionar_comentario/<demanda_id>', methods=['POST'])
@login_required
def adicionar_comentario(demanda_id):
    conn = get_db()
    autor = session['user_nome']
    comentario = request.form['comentario']
    conn.execute(
        "INSERT INTO comentarios (demanda_id, comentario, autor, data) VALUES (?, ?, ?, ?)",
        (demanda_id, comentario, autor, datetime.now().strftime("%d/%m/%Y %H:%M"))
    )
    conn.commit()
    conn.close()

    log_action(
        session.get('user_nome', 'sistema'),
        'ADICIONAR_COMENTARIO',
        f'Comentário na demanda id={demanda_id}',
        request.remote_addr
    )

    return redirect(f'/detalhes/{demanda_id}')


@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db()

    f_de          = request.args.get('de', '').strip()
    f_ate         = request.args.get('ate', '').strip()
    f_responsavel = request.args.get('responsavel', '').strip()
    f_prioridade  = request.args.get('prioridade', '').strip()
    f_status      = request.args.get('status', '').strip()

    conditions = []
    params = []

    if f_de:
        conditions.append("substr(d.data_criacao,7,4)||'-'||substr(d.data_criacao,4,2)||'-'||substr(d.data_criacao,1,2) >= ?")
        params.append(f_de)
    if f_ate:
        conditions.append("substr(d.data_criacao,7,4)||'-'||substr(d.data_criacao,4,2)||'-'||substr(d.data_criacao,1,2) <= ?")
        params.append(f_ate)
    if f_responsavel:
        conditions.append("d.solicitante = ?")
        params.append(f_responsavel)
    if f_prioridade:
        conditions.append("p.valor = ?")
        params.append(f_prioridade)
    if f_status:
        conditions.append("d.status = ?")
        params.append(f_status)

    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    and_or_where = "AND" if where else "WHERE"

    base_query = f"""
        SELECT d.*, p.valor as prioridade_nome
        FROM demandas d
        LEFT JOIN prioridades p ON d.id_prioridade = p.id
        {where}
    """

    total      = conn.execute(f"SELECT COUNT(*) FROM demandas d LEFT JOIN prioridades p ON d.id_prioridade = p.id {where}", params).fetchone()[0]
    abertas    = conn.execute(f"SELECT COUNT(*) FROM demandas d LEFT JOIN prioridades p ON d.id_prioridade = p.id {where} {and_or_where} d.status = 'Aberta'", params).fetchone()[0]
    concluidas = conn.execute(f"SELECT COUNT(*) FROM demandas d LEFT JOIN prioridades p ON d.id_prioridade = p.id {where} {and_or_where} d.status = 'Concluída'", params).fetchone()[0]

    atrasadas_count = conn.execute(
        f"SELECT COUNT(*) FROM demandas d LEFT JOIN prioridades p ON d.id_prioridade = p.id {where} {and_or_where} d.status = 'Aberta' AND d.prazo IS NOT NULL AND d.prazo < date('now')", params
    ).fetchone()[0]

    atrasadas_list = conn.execute(
        f"{base_query} {and_or_where} d.status = 'Aberta' AND d.prazo IS NOT NULL AND d.prazo < date('now') ORDER BY d.prazo ASC", params
    ).fetchall()

    criticas = conn.execute(
        f"{base_query} {and_or_where} p.valor = 'Alta' AND d.status = 'Aberta' ORDER BY d.data_criacao ASC", params
    ).fetchall()

    por_responsavel = conn.execute(
        f"SELECT d.solicitante, COUNT(*) as total FROM demandas d LEFT JOIN prioridades p ON d.id_prioridade = p.id {where} {and_or_where} d.status = 'Aberta' GROUP BY d.solicitante ORDER BY total DESC", params
    ).fetchall()

    tempo_medio = conn.execute(f"""
        SELECT AVG(
            julianday(substr(d.data_criacao,7,4)||'-'||substr(d.data_criacao,4,2)||'-'||substr(d.data_criacao,1,2)) -
            julianday(substr(d.data_criacao,7,4)||'-'||substr(d.data_criacao,4,2)||'-'||substr(d.data_criacao,1,2))
        ) FROM demandas d LEFT JOIN prioridades p ON d.id_prioridade = p.id
        {where} {and_or_where} d.status = 'Concluída'
    """, params).fetchone()[0]

    responsaveis = conn.execute("SELECT DISTINCT solicitante FROM demandas WHERE solicitante IS NOT NULL ORDER BY solicitante").fetchall()

    conn.close()

    pct_abertas    = round((abertas / total * 100), 1) if total else 0
    pct_concluidas = round((concluidas / total * 100), 1) if total else 0
    pct_atrasadas  = round((atrasadas_count / total * 100), 1) if total else 0

    return render_template('dashboard.html',
        total=total,
        abertas=abertas,
        concluidas=concluidas,
        atrasadas=atrasadas_count,
        atrasadas_list=atrasadas_list,
        criticas=criticas,
        por_responsavel=por_responsavel,
        tempo_medio=round(tempo_medio, 1) if tempo_medio else 0,
        pct_abertas=pct_abertas,
        pct_concluidas=pct_concluidas,
        pct_atrasadas=pct_atrasadas,
        responsaveis=responsaveis,
        f_de=f_de,
        f_ate=f_ate,
        f_responsavel=f_responsavel,
        f_prioridade=f_prioridade,
        f_status=f_status,
    )


@app.route('/logs')
@login_required
def logs():
    conn = get_db()
    entries = conn.execute(
        'SELECT * FROM logs ORDER BY id DESC LIMIT 200'
    ).fetchall()
    conn.close()
    return render_template('logs.html', logs=entries)


@app.route('/exportar/atrasadas')
@login_required
def exportar_atrasadas():
    conn = get_db()
    atrasadas = conn.execute("""
        SELECT d.titulo, d.solicitante, d.prazo, d.data_criacao, p.valor as prioridade
        FROM demandas d
        LEFT JOIN prioridades p ON d.id_prioridade = p.id
        WHERE d.status = 'Aberta' AND d.prazo IS NOT NULL AND d.prazo < date('now')
        ORDER BY d.prazo ASC
    """).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Demandas Atrasadas"

    ws.merge_cells('A1:E1')
    ws['A1'] = f"Relatório de Demandas Atrasadas — {date.today().strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['Título', 'Solicitante', 'Prazo', 'Data de Criação', 'Prioridade']
    header_fill = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    red_fill = PatternFill(start_color='FCEBEB', end_color='FCEBEB', fill_type='solid')
    for row_idx, d in enumerate(atrasadas, 4):
        ws.cell(row=row_idx, column=1, value=d['titulo'])
        ws.cell(row=row_idx, column=2, value=d['solicitante'])
        ws.cell(row=row_idx, column=3, value=d['prazo'])
        ws.cell(row=row_idx, column=4, value=d['data_criacao'])
        ws.cell(row=row_idx, column=5, value=d['prioridade'])
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).fill = red_fill

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"atrasadas_{date.today().strftime('%Y%m%d')}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')